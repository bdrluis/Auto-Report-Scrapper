import tkinter as tk
from tkinter import messagebox
import os
import openpyxl
from bs4 import BeautifulSoup
import pandas as pd
import openpyxl.styles as styles

def read_html(file_path):
    with open(file_path, 'r', encoding='utf-8') as file:
        html_content = file.read()
    return BeautifulSoup(html_content, 'html.parser')

def extract_data(soup):
    target_td = soup.find('td', title="Commission + Swap + Profit + Taxes")
    ticket_td = soup.find('td', string='Ticket')

    extracted_data = []
    if target_td and ticket_td:
        extracted_data.append(ticket_td.get_text(strip=True))
        filtered_td_elements = ticket_td.find_all_next('td')

        for current_td in filtered_td_elements:
            extracted_data.append(current_td.get_text(strip=True))
            if current_td == target_td:
                break

    return extracted_data[:-7]

def filter_data(data):
    balance_indices = [index for index, value in enumerate(data) if value == 'balance']
    cancelled_indices = [index for index, value in enumerate(data) if value == 'cancelled']
    indices_to_drop = [idx for balance_idx in balance_indices for idx in range(balance_idx - 2, balance_idx + 3)] + \
                      [idx for cancelled_idx in cancelled_indices for idx in range(cancelled_idx - 10, cancelled_idx + 1)]
    return [value for i, value in enumerate(data) if i not in indices_to_drop]

def main():
    try:
        file_name = input("Enter the name of the HTML file: ")
        file_path = rf'C:\Users\GAMERXXX\Desktop\stats\{file_name}.htm'

        if not os.path.isfile(file_path):
            raise FileNotFoundError(f"The file '{file_path}' does not exist.")

        soup = read_html(file_path)
        data = extract_data(soup)
        filtered_data = filter_data(data)

        rows = [filtered_data[i:i+14] for i in range(0, len(filtered_data), 14)]

        # Define column names
        columns = rows[0]
        
        # Create a DataFrame excluding the first row (column names)
        df = pd.DataFrame(rows[1:], columns=columns)
        
        
        # some formatting
        # split the open time and close time columns into date and datetime
        df[['Open Day', 'Open Time']] = df['Open Time'].str.split(' ', n=1, expand=True)
        df[['Close Day', 'Close Time']] = df['Close Time'].str.split(' ', n=1, expand=True)
        
        
        # delete some useless columns
        columns_drop= ['Ticket','Commission','Taxes', 'Swap','Size','Close Day']
        df.drop(columns_drop ,axis = 1, inplace=True)
        
        # rename second price to close_price
        index_of_second_price = df.columns.to_list().index("Price", df.columns.to_list().index("Price") + 1)
        df.columns.values[index_of_second_price] = 'close_price'
        
        #re arrange columns
        df = df[[ 'Open Day', 'Open Time','Close Time', 'Type', 'Item','Price','S / L', 'T / P','close_price','Profit']]
        
        
        
        numeric_columns = ['Price', 'S / L', 'T / P', 'close_price', 'Profit']
        df[numeric_columns] = df[numeric_columns].astype(float)
        str_columns = ['Type', 'Item']
        df[str_columns] = df[str_columns].astype(str)
        
        df['Open Day'] = df['Open Day'].str.replace('.', '-')
        
        # Convert 'Open Day' to datetime with the specified format
        df['Open Day'] = pd.to_datetime(df['Open Day'], format='%Y-%m-%d')
        
        # Convert 'Open Time' and 'Close Time' to time directly
        df['Open Time'] = pd.to_datetime(df['Open Time'], format='%H:%M:%S')
        df['Close Time'] = pd.to_datetime(df['Close Time'], format='%H:%M:%S')
        # data is successfuly exctracted and cleaned by now , you can check the df
        
        excel_file = r'C:\Users\GAMERXXX\Desktop\stats\trading1.xlsx'
        sheet_name = 'Sheet1'  # Replace with the actual sheet name
        
        # Load the existing/target Excel file
        existing_data = pd.read_excel(excel_file, sheet_name=sheet_name)
        
        # Determine the starting row for the new data
        start_row = existing_data.shape[0] + 1  # Add 1 to skip one full row 
        with pd.ExcelWriter(excel_file, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            
            # Append the DataFrame to the specified sheet starting from the start_row
            df.to_excel(writer, sheet_name, index=False, header=False, startrow=start_row)
        
        # Open the Excel file to adjust the formatting
        workbook = openpyxl.load_workbook(excel_file)
        sheet = workbook[sheet_name]
        
        # Retrieve formatting information from the existing sheet
        formatting_info = {}
        for col_idx, column in enumerate(df.columns, start=1):
            cell = sheet.cell(row=2, column=col_idx)  # Assuming row 2 has data with formatting
            formatting_info[column] = {
                'alignment': styles.Alignment(**cell.alignment.__dict__),
                'font': styles.Font(**cell.font.__dict__),
                'number_format': cell.number_format,
            }
        
        # Apply the formatting information to the entire sheet
        for col_idx, column in enumerate(df.columns, start=1):
            for row_idx in range(1, sheet.max_row + 1):
                cell = sheet.cell(row=row_idx, column=col_idx)
                cell.alignment = formatting_info[column]['alignment']
                cell.font = formatting_info[column]['font']
                cell.number_format = formatting_info[column]['number_format']
        
        # Save the changes
        workbook.save(excel_file)

        # Display success message using GUI
        root = tk.Tk()
        root.withdraw()
        messagebox.showinfo("Success", f"Data has been modified successfully in {excel_file}")

    except FileNotFoundError as e:
        print(f"Error: {e}")
    except Exception as e:
        print(f"Error: {e}")

if __name__ == "__main__":
    main()
